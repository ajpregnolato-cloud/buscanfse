import json
import re
import subprocess
import time
from datetime import datetime
from pathlib import Path
from typing import Any
import tkinter as tk
from tkinter import filedialog, messagebox

import customtkinter as ctk
import pandas as pd
import requests
import smtplib
from email.message import EmailMessage
from email.utils import formatdate

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


DANFSE_PROD = "https://adn.nfse.gov.br/danfse"
APP_NAME = "NFS-e | DANFSE em Lote (Colar ou Excel)"
CONFIG_FILE = "config.json"


ctk.set_appearance_mode("System")
ctk.set_default_color_theme("blue")


def normalize_key(s: str) -> str:
    return "".join(ch for ch in (s or "").strip() if ch.isdigit())


def is_pdf_response(resp: requests.Response) -> bool:
    ct = (resp.headers.get("content-type") or "").lower()
    return ("application/pdf" in ct) or (resp.content[:4] == b"%PDF")


def get_with_retry(url: str, cert_tuple, timeout: int, retries: int = 4):
    last = None
    for attempt in range(1, retries + 1):
        try:
            r = requests.get(
                url,
                cert=cert_tuple,
                timeout=timeout,
                headers={
                    "Accept": "application/pdf, application/json;q=0.9, */*;q=0.8",
                    "User-Agent": "NfseClient/1.0 (Python requests)",
                },
            )
            last = r
            if r.status_code in (502, 503, 504, 520, 521, 522):
                time.sleep(1.5 * attempt)
                continue
            return r
        except requests.RequestException as e:
            last = e
            time.sleep(1.5 * attempt)
    return last


def safe_int(value: Any, default: int) -> int:
    try:
        return int(str(value).strip())
    except (TypeError, ValueError):
        return default


def safe_float(value: Any, default: float) -> float:
    try:
        return float(str(value).strip())
    except (TypeError, ValueError):
        return default


def dedup_keep_order(values: list[str]) -> list[str]:
    seen = set()
    uniq = []
    for value in values:
        if value not in seen:
            uniq.append(value)
            seen.add(value)
    return uniq




def extract_cert_identity(cert_path: Path) -> tuple[str, str]:
    if not cert_path.exists():
        return ("Não identificado", "Não identificado")

    def only_digits(value: str) -> str:
        return "".join(ch for ch in (value or "") if ch.isdigit())

    def format_cnpj(value: str) -> str:
        digits = only_digits(value)
        if len(digits) >= 14:
            c = digits[-14:]
            return f"{c[:2]}.{c[2:5]}.{c[5:8]}/{c[8:12]}-{c[12:14]}"
        return "Não identificado"

    def decode_possible_hex(value: str) -> str:
        cleaned = value.strip().replace(":", "").replace(" ", "")
        if len(cleaned) < 2 or len(cleaned) % 2 != 0:
            return ""
        if not re.fullmatch(r"[0-9A-Fa-f]+", cleaned):
            return ""
        try:
            return bytes.fromhex(cleaned).decode("latin-1", errors="ignore")
        except ValueError:
            return ""

    def run_openssl(args: list[str]) -> str:
        try:
            r = subprocess.run(args, check=True, capture_output=True, text=True)
            return (r.stdout or "")
        except (FileNotFoundError, subprocess.CalledProcessError):
            return ""

    subject_rfc2253 = run_openssl([
        "openssl", "x509", "-in", str(cert_path), "-noout", "-subject", "-nameopt", "RFC2253,utf8"
    ]).strip()
    if "subject=" in subject_rfc2253:
        subject_rfc2253 = subject_rfc2253.split("subject=", 1)[1].strip()

    subject_default = run_openssl([
        "openssl", "x509", "-in", str(cert_path), "-noout", "-subject"
    ]).strip()

    cert_text = run_openssl([
        "openssl", "x509", "-in", str(cert_path), "-noout", "-text"
    ])

    subject_joined = f"{subject_rfc2253}\n{subject_default}"
    haystack = f"{subject_joined}\n{cert_text}"

    social_name = "Não identificado"
    cnpj = "Não identificado"

    # Razão Social (prioriza O=, fallback CN=)
    o_match = re.search(r"(?:^|,|/)\s*O\s*=\s*([^,\/\n]+)", subject_joined, flags=re.IGNORECASE)
    cn_match = re.search(r"(?:^|,|/)\s*CN\s*=\s*([^,\/\n]+)", subject_joined, flags=re.IGNORECASE)
    if o_match:
        social_name = o_match.group(1).strip()
    elif cn_match:
        social_name = cn_match.group(1).strip()

    # 1) serialNumber do subject
    serial_match = re.search(r"serialNumber\s*=\s*([^,\/\n]+)", subject_joined, flags=re.IGNORECASE)
    if serial_match:
        cnpj = format_cnpj(serial_match.group(1))

    # 2) OID ICP-Brasil do CNPJ: 2.16.76.1.3.3 em formatos variados
    if cnpj == "Não identificado":
        oid_patterns = [
            r"(?:OID\.)?2\.16\.76\.1\.3\.3\s*[:=]\s*([^\n,]+)",
            r"(?:OID\.)?2\.16\.76\.1\.3\.3\s*:\s*<[^>]+>\s*:?\s*([^\n,]+)",
            r"othername\s*:?\s*(?:OID\.)?2\.16\.76\.1\.3\.3\s*[:=]\s*([^\n,]+)",
            r"CNPJ\s*[:=]\s*([\d\.\-/]{14,})",
        ]

        candidates: list[str] = []
        for pattern in oid_patterns:
            candidates.extend(re.findall(pattern, haystack, flags=re.IGNORECASE))

        for candidate in candidates:
            candidate = re.sub(r"<(?:ASN1_)?(?:UTF8STRING|PRINTABLESTRING|IA5STRING|OCTET_STRING)>", "", candidate, flags=re.IGNORECASE)
            candidate = candidate.strip()

            direct = format_cnpj(candidate)
            if direct != "Não identificado":
                cnpj = direct
                break

            decoded = decode_possible_hex(candidate)
            if decoded:
                from_hex = format_cnpj(decoded)
                if from_hex != "Não identificado":
                    cnpj = from_hex
                    break

    # 3) fallback conservador: qualquer bloco com 14+ dígitos no texto completo
    if cnpj == "Não identificado":
        for chunk in re.findall(r"[\d\.\-/]{14,}", haystack):
            candidate = format_cnpj(chunk)
            if candidate != "Não identificado":
                cnpj = candidate
                break

    return (social_name, cnpj)


def send_email_smtp(
    smtp_host: str,
    smtp_port: int,
    smtp_user: str,
    smtp_pass: str,
    from_email: str,
    to_email: str,
    subject: str,
    body: str,
    attachment_path: Path,
):
    msg = EmailMessage()
    msg["From"] = from_email
    msg["To"] = to_email
    msg["Date"] = formatdate(localtime=True)
    msg["Subject"] = subject
    msg.set_content(body)

    data = attachment_path.read_bytes()
    msg.add_attachment(data, maintype="application", subtype="pdf", filename=attachment_path.name)

    with smtplib.SMTP(smtp_host, smtp_port, timeout=30) as s:
        s.ehlo()
        s.starttls()
        s.ehlo()
        s.login(smtp_user, smtp_pass)
        s.send_message(msg)


def auto_fit_excel_columns(path: Path, widths: list[int]):
    wb = load_workbook(path)
    ws = wb.active
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    wb.save(path)


def create_excel_template(path: Path):
    df = pd.DataFrame([{"CHAVE": "", "EMAIL": "", "ASSUNTO": "", "CORPO": ""}])
    df.to_excel(path, index=False, sheet_name="DANFSE")
    auto_fit_excel_columns(path, [48, 30, 40, 60])


def read_excel_rows(path: Path):
    df = pd.read_excel(path, dtype=str).fillna("")
    columns = {str(c).strip().upper(): c for c in df.columns}

    if "CHAVE" not in columns:
        raise RuntimeError("Planilha inválida: coluna CHAVE é obrigatória.")

    def get_col(name: str) -> str:
        return columns[name] if name in columns else ""

    key_col = get_col("CHAVE")
    email_col = get_col("EMAIL")
    assunto_col = get_col("ASSUNTO")
    corpo_col = get_col("CORPO")

    rows = []
    for _, row in df.iterrows():
        chave = normalize_key(str(row.get(key_col, "")))
        if not chave:
            continue

        rows.append(
            {
                "chave": chave,
                "email": str(row.get(email_col, "")).strip() if email_col else "",
                "assunto": str(row.get(assunto_col, "")).strip() if assunto_col else "",
                "corpo": str(row.get(corpo_col, "")).strip() if corpo_col else "",
            }
        )

    uniq_keys = dedup_keep_order([item["chave"] for item in rows])
    row_by_key = {item["chave"]: item for item in rows}
    return [row_by_key[k] for k in uniq_keys]


def now_stamp() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S")


def salvar_falhas_txt(falhas: list[str], pasta: Path) -> Path:
    p = pasta / f"falhas_{now_stamp()}.txt"
    p.write_text("\n".join(falhas), encoding="utf-8")
    return p


def salvar_resultado_xlsx(linhas: list[dict], pasta: Path) -> Path:
    p = pasta / f"resultado_{now_stamp()}.xlsx"
    cols = ["chave", "status", "pdf_path", "email_to", "erro", "origem"]
    df = pd.DataFrame(linhas or [], columns=cols)
    df.columns = ["CHAVE", "STATUS", "PDF_PATH", "EMAIL_TO", "ERRO", "ORIGEM"]
    df.to_excel(p, index=False, sheet_name="RESULTADO")
    auto_fit_excel_columns(p, [48, 12, 60, 30, 60, 18])
    return p


class App(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title(APP_NAME)
        self.geometry("1160x780")
        self.resizable(False, False)

        self.cfg = self.load_config()

        self.cert_pem = tk.StringVar(value=self.cfg.get("cert_pem", str(Path.cwd() / "cert_out" / "client_cert.pem")))
        self.key_pem = tk.StringVar(value=self.cfg.get("key_pem", str(Path.cwd() / "cert_out" / "client_key.pem")))
        self.pfx_file = tk.StringVar(value=self.cfg.get("pfx_file", ""))
        self.pfx_password = tk.StringVar(value=self.cfg.get("pfx_password", ""))
        self.pfx_output_dir = tk.StringVar(value=self.cfg.get("pfx_output_dir", str(Path.cwd() / "cert_out")))

        self.save_dir = tk.StringVar(value=self.cfg.get("save_dir", str(Path.cwd() / "downloads")))

        self.smtp_host = tk.StringVar(value=self.cfg.get("smtp_host", ""))
        self.smtp_port = tk.StringVar(value=str(self.cfg.get("smtp_port", "587")))
        self.smtp_user = tk.StringVar(value=self.cfg.get("smtp_user", ""))
        self.smtp_pass = tk.StringVar(value=self.cfg.get("smtp_pass", ""))
        self.from_email = tk.StringVar(value=self.cfg.get("from_email", ""))
        self.default_to = tk.StringVar(value=self.cfg.get("default_to", ""))
        self.default_subject = tk.StringVar(value=self.cfg.get("default_subject", "NFSe - DANFSE (PDF)"))
        self.default_body = tk.StringVar(value=self.cfg.get("default_body", "Segue em anexo o DANFSE (PDF)."))

        self.send_email_each = tk.BooleanVar(value=bool(self.cfg.get("send_email_each", False)))
        self.progress = tk.DoubleVar(value=0.0)

        self.pause_between_items = tk.DoubleVar(value=float(self.cfg.get("pause_between_items", 0.5)))
        self.reprocess_rounds = tk.IntVar(value=int(self.cfg.get("reprocess_rounds", 3)))
        self.cooldown_between_rounds = tk.IntVar(value=int(self.cfg.get("cooldown_between_rounds", 10)))
        self.cooldown_every_n = tk.IntVar(value=int(self.cfg.get("cooldown_every_n", 5)))
        self.cooldown_seconds = tk.IntVar(value=int(self.cfg.get("cooldown_seconds", 3)))

        self.excel_rows = []
        self.cert_social_name = tk.StringVar(value="Não identificado")
        self.cert_cnpj = tk.StringVar(value="Não identificado")

        self._build_ui()
        self.refresh_certificate_identity()

    def load_config(self):
        p = Path(CONFIG_FILE)
        if p.exists():
            try:
                return json.loads(p.read_text(encoding="utf-8"))
            except Exception:
                return {}
        return {}

    def save_config(self):
        cfg = {
            "cert_pem": self.cert_pem.get().strip(),
            "key_pem": self.key_pem.get().strip(),
            "pfx_file": self.pfx_file.get().strip(),
            "pfx_password": self.pfx_password.get(),
            "pfx_output_dir": self.pfx_output_dir.get().strip(),
            "save_dir": self.save_dir.get().strip(),
            "smtp_host": self.smtp_host.get().strip(),
            "smtp_port": safe_int(self.smtp_port.get(), 587),
            "smtp_user": self.smtp_user.get().strip(),
            "smtp_pass": self.smtp_pass.get(),
            "from_email": self.from_email.get().strip(),
            "default_to": self.default_to.get().strip(),
            "default_subject": self.default_subject.get().strip(),
            "default_body": self.body_txt.get("1.0", "end").strip(),
            "send_email_each": bool(self.send_email_each.get()),
            "pause_between_items": safe_float(self.pause_between_items.get(), 0.5),
            "reprocess_rounds": safe_int(self.reprocess_rounds.get(), 3),
            "cooldown_between_rounds": safe_int(self.cooldown_between_rounds.get(), 10),
            "cooldown_every_n": safe_int(self.cooldown_every_n.get(), 5),
            "cooldown_seconds": safe_int(self.cooldown_seconds.get(), 3),
        }
        Path(CONFIG_FILE).write_text(json.dumps(cfg, ensure_ascii=False, indent=2), encoding="utf-8")
        self.cfg = cfg

    def log(self, msg: str):
        self.txt_log.insert("end", msg + "\n")
        self.txt_log.see("end")
        self.update_idletasks()

    def _cert_tuple(self):
        cert = Path(self.cert_pem.get().strip().strip('"'))
        key = Path(self.key_pem.get().strip().strip('"'))
        if not cert.exists():
            raise FileNotFoundError(f"client_cert.pem não encontrado: {cert}")
        if not key.exists():
            raise FileNotFoundError(f"client_key.pem não encontrado: {key}")
        return (str(cert), str(key))

    def _out_dir(self):
        d = Path(self.save_dir.get().strip().strip('"'))
        d.mkdir(parents=True, exist_ok=True)
        return d

    def smtp_ready(self):
        host = self.smtp_host.get().strip()
        user = self.smtp_user.get().strip()
        pw = self.smtp_pass.get()
        to_e = self.default_to.get().strip()
        from_e = self.from_email.get().strip() or user
        return bool(host and user and pw and to_e and from_e)

    def baixar_pdf(self, chave: str) -> tuple[Path | None, str]:
        url = f"{DANFSE_PROD}/{chave}"
        resp = get_with_retry(url, self._cert_tuple(), timeout=70, retries=4)

        if isinstance(resp, Exception):
            return None, f"rede: {resp}"

        if resp.status_code == 200 and is_pdf_response(resp):
            out = self._out_dir() / f"DANFSE_{chave}.pdf"
            out.write_bytes(resp.content)
            return out, ""

        ct = resp.headers.get("content-type", "")
        err_txt = f"status={resp.status_code} ct={ct}"
        try:
            snippet = resp.text[:200].replace("\n", " ").strip()
            if snippet:
                err_txt += f" resp={snippet}"
        except Exception:
            pass
        return None, err_txt

    def send_email_for_pdf(self, chave: str, pdf: Path, email: str = "", assunto: str = "", corpo: str = ""):
        host = self.smtp_host.get().strip()
        port = safe_int(self.smtp_port.get(), 587)
        user = self.smtp_user.get().strip()
        pw = self.smtp_pass.get()
        from_e = self.from_email.get().strip() or user

        to_e = (email or "").strip() or self.default_to.get().strip()
        subject = (assunto or "").strip() or self.default_subject.get().strip()
        body = (corpo or "").strip() or self.body_txt.get("1.0", "end").strip()

        subject = f"{subject} | {chave}"
        send_email_smtp(host, port, user, pw, from_e, to_e, subject, body, pdf)
        return to_e

    def _build_ui(self):
        tabview = ctk.CTkTabview(self, width=1120, height=740)
        tabview.pack(fill="both", expand=True, padx=12, pady=12)

        tab_lote = tabview.add("Lote")
        tab_cfg = tabview.add("Configuração")

        ctk.CTkLabel(tab_lote, text="Busca NFSe - Grupo Supply Service", font=ctk.CTkFont(size=20, weight="bold")).pack(anchor="w", pady=(10, 2), padx=10)
        ctk.CTkLabel(tab_lote, textvariable=self.cert_social_name, font=ctk.CTkFont(size=14, weight="bold")).pack(anchor="w", padx=10)
        ctk.CTkLabel(tab_lote, textvariable=self.cert_cnpj, font=ctk.CTkFont(size=13)).pack(anchor="w", padx=10, pady=(0, 8))

        ctk.CTkLabel(tab_lote, text="Opção 1 — Colar chaves (1 por linha):", font=ctk.CTkFont(size=14, weight="bold")).pack(anchor="w", pady=(8, 6), padx=10)
        self.txt_keys = ctk.CTkTextbox(tab_lote, width=1080, height=180)
        self.txt_keys.pack(anchor="w", padx=10)

        rowbtn = ctk.CTkFrame(tab_lote, fg_color="transparent")
        rowbtn.pack(anchor="w", padx=10, pady=(8, 8))
        ctk.CTkButton(rowbtn, text="Limpar", command=lambda: self.txt_keys.delete("1.0", "end"), width=120).pack(side="left")
        ctk.CTkButton(rowbtn, text="Executar lote (texto)", command=self.run_from_text, width=170).pack(side="left", padx=8)
        ctk.CTkButton(rowbtn, text="Baixar modelo Excel", command=self.download_template, width=170).pack(side="left", padx=8)
        ctk.CTkButton(rowbtn, text="Importar Excel", command=self.import_excel, width=140).pack(side="left", padx=8)
        ctk.CTkButton(rowbtn, text="Executar lote (Excel)", command=self.run_from_excel, width=170).pack(side="left", padx=8)

        self.lbl_excel = ctk.CTkLabel(tab_lote, text="Nenhum Excel importado.")
        self.lbl_excel.pack(anchor="w", padx=10, pady=(0, 8))

        self.chk_email = ctk.CTkCheckBox(tab_lote, text="Enviar e-mail automaticamente para cada PDF", variable=self.send_email_each)
        self.chk_email.pack(anchor="w", padx=10, pady=(0, 8))

        self.pb = ctk.CTkProgressBar(tab_lote, width=1080)
        self.pb.pack(anchor="w", padx=10)
        self.pb.set(0)

        self.txt_log = ctk.CTkTextbox(tab_lote, width=1080, height=280)
        self.txt_log.pack(anchor="w", padx=10, pady=(8, 8))

        ctk.CTkLabel(tab_lote, text="Todos os Direitos Reservados - Versão 02/2026.", font=ctk.CTkFont(size=12)).pack(anchor="w", padx=10, pady=(0, 6))

        self._build_config_tab(tab_cfg)

    def _build_config_tab(self, tab_cfg):
        container = ctk.CTkScrollableFrame(tab_cfg, width=1100, height=700)
        container.pack(fill="both", expand=True, padx=10, pady=10)

        ctk.CTkLabel(container, text="Certificado mTLS", font=ctk.CTkFont(size=16, weight="bold")).pack(anchor="w", pady=(8, 6))

        self._labeled_entry(container, "client_cert.pem", self.cert_pem, self.pick_cert)
        self._labeled_entry(container, "client_key.pem", self.key_pem, self.pick_key)
        self._labeled_entry(container, "Pasta para salvar PDFs", self.save_dir, self.pick_dir)

        ctk.CTkLabel(container, text="Conversão PFX → PEM (troca de CNPJ)", font=ctk.CTkFont(size=16, weight="bold")).pack(anchor="w", pady=(16, 6))
        self._labeled_entry(container, "Arquivo .pfx/.p12", self.pfx_file, self.pick_pfx)
        self._labeled_entry(container, "Senha do PFX", self.pfx_password, None, show="*")
        self._labeled_entry(container, "Pasta de saída dos PEM", self.pfx_output_dir, self.pick_pfx_out_dir)

        ctk.CTkButton(container, text="Converter PFX e substituir certificados ativos", command=self.convert_pfx_to_pem, width=360).pack(anchor="w", pady=(6, 8))

        ctk.CTkLabel(container, text="Robustez do lote", font=ctk.CTkFont(size=16, weight="bold")).pack(anchor="w", pady=(8, 6))
        self._labeled_entry(container, "Pausa entre itens (s)", self.pause_between_items)
        self._labeled_entry(container, "Rodadas de reprocessamento", self.reprocess_rounds)
        self._labeled_entry(container, "Cooldown entre rodadas (s)", self.cooldown_between_rounds)
        self._labeled_entry(container, "Cooldown a cada N itens", self.cooldown_every_n)
        self._labeled_entry(container, "Cooldown (s)", self.cooldown_seconds)

        ctk.CTkLabel(container, text="SMTP", font=ctk.CTkFont(size=16, weight="bold")).pack(anchor="w", pady=(16, 6))
        self._labeled_entry(container, "SMTP Host", self.smtp_host)
        self._labeled_entry(container, "SMTP Port", self.smtp_port)
        self._labeled_entry(container, "SMTP User", self.smtp_user)
        self._labeled_entry(container, "SMTP Pass", self.smtp_pass, None, show="*")
        self._labeled_entry(container, "From", self.from_email)
        self._labeled_entry(container, "To padrão", self.default_to)
        self._labeled_entry(container, "Assunto padrão", self.default_subject)

        ctk.CTkLabel(container, text="Corpo padrão").pack(anchor="w")
        self.body_txt = ctk.CTkTextbox(container, width=900, height=120)
        self.body_txt.pack(anchor="w", pady=(4, 8))
        self.body_txt.insert("1.0", self.default_body.get())

        ctk.CTkButton(container, text="Salvar configuração", command=self.on_save_config, width=220).pack(anchor="w", pady=(6, 20))

    def _labeled_entry(self, parent, label, var, picker=None, show=None):
        ctk.CTkLabel(parent, text=label).pack(anchor="w", pady=(4, 2))
        row = ctk.CTkFrame(parent, fg_color="transparent")
        row.pack(anchor="w", fill="x")
        ctk.CTkEntry(row, textvariable=var, width=760, show=show).pack(side="left", padx=(0, 6))
        if picker:
            ctk.CTkButton(row, text="Selecionar...", width=120, command=picker).pack(side="left")

    def on_save_config(self):
        self.save_config()
        self.refresh_certificate_identity()
        messagebox.showinfo("OK", f"Configuração salva em {Path(CONFIG_FILE).resolve()}")

    def pick_cert(self):
        p = filedialog.askopenfilename(title="Selecione client_cert.pem", filetypes=[("PEM", "*.pem"), ("Todos", "*.*")])
        if p:
            self.cert_pem.set(p)
            self.refresh_certificate_identity()

    def pick_key(self):
        p = filedialog.askopenfilename(title="Selecione client_key.pem", filetypes=[("PEM", "*.pem"), ("Todos", "*.*")])
        if p:
            self.key_pem.set(p)

    def pick_dir(self):
        p = filedialog.askdirectory(title="Selecione a pasta de saída")
        if p:
            self.save_dir.set(p)

    def pick_pfx(self):
        p = filedialog.askopenfilename(title="Selecione arquivo PFX/P12", filetypes=[("Certificado", "*.pfx *.p12"), ("Todos", "*.*")])
        if p:
            self.pfx_file.set(p)

    def pick_pfx_out_dir(self):
        p = filedialog.askdirectory(title="Selecione a pasta para gerar client_cert.pem/client_key.pem")
        if p:
            self.pfx_output_dir.set(p)

    def convert_pfx_to_pem(self):
        pfx_path = Path(self.pfx_file.get().strip())
        pfx_pass = self.pfx_password.get()
        out_dir = Path(self.pfx_output_dir.get().strip() or Path.cwd() / "cert_out")

        if not pfx_path.exists():
            messagebox.showerror("Erro", "Arquivo PFX/P12 não encontrado.")
            return

        out_dir.mkdir(parents=True, exist_ok=True)
        cert_out = out_dir / "client_cert.pem"
        key_out = out_dir / "client_key.pem"

        cmd_cert = [
            "openssl",
            "pkcs12",
            "-in",
            str(pfx_path),
            "-clcerts",
            "-nokeys",
            "-out",
            str(cert_out),
            "-passin",
            f"pass:{pfx_pass}",
        ]
        cmd_key = [
            "openssl",
            "pkcs12",
            "-in",
            str(pfx_path),
            "-nocerts",
            "-nodes",
            "-out",
            str(key_out),
            "-passin",
            f"pass:{pfx_pass}",
        ]

        try:
            subprocess.run(cmd_cert, check=True, capture_output=True, text=True)
            subprocess.run(cmd_key, check=True, capture_output=True, text=True)
        except FileNotFoundError:
            messagebox.showerror("Erro", "OpenSSL não encontrado no sistema. Instale e tente novamente.")
            return
        except subprocess.CalledProcessError as e:
            err = (e.stderr or e.stdout or "").strip()[:300]
            messagebox.showerror("Erro", f"Falha ao converter PFX:\n{err}")
            return

        self.cert_pem.set(str(cert_out))
        self.key_pem.set(str(key_out))
        self.refresh_certificate_identity()
        self.save_config()
        messagebox.showinfo("Sucesso", f"Conversão concluída!\n\nCert: {cert_out}\nKey: {key_out}\n\nOs arquivos ativos foram atualizados.")

    def refresh_certificate_identity(self):
        cert_path = Path(self.cert_pem.get().strip().strip('"'))
        razao, cnpj = extract_cert_identity(cert_path)
        self.cert_social_name.set(f"Razão Social: {razao}")
        self.cert_cnpj.set(f"CNPJ do Certificado: {cnpj}")

    def download_template(self):
        path = filedialog.asksaveasfilename(title="Salvar planilha modelo", defaultextension=".xlsx", filetypes=[("Excel", "*.xlsx")])
        if not path:
            return
        p = Path(path)
        create_excel_template(p)
        messagebox.showinfo("OK", f"Planilha modelo salva em:\n{p}")

    def import_excel(self):
        p = filedialog.askopenfilename(title="Importar planilha Excel", filetypes=[("Excel", "*.xlsx")])
        if not p:
            return
        try:
            rows = read_excel_rows(Path(p))
        except Exception as e:
            messagebox.showerror("Erro", f"Falha ao importar Excel:\n{e}")
            return
        self.excel_rows = rows
        self.lbl_excel.configure(text=f"Excel importado: {Path(p).name} | Linhas válidas: {len(rows)}")
        messagebox.showinfo("OK", f"Importado com sucesso.\nLinhas válidas: {len(rows)}")

    def run_from_text(self):
        keys = []
        for line in self.txt_keys.get("1.0", "end").splitlines():
            k = normalize_key(line)
            if k:
                keys.append(k)

        uniq = dedup_keep_order(keys)
        if not uniq:
            messagebox.showerror("Erro", "Nenhuma chave válida no texto.")
            return

        self.save_config()
        items = [{"chave": k, "email": "", "assunto": "", "corpo": "", "origem": "texto"} for k in uniq]
        self.run_batch_items(items)

    def run_from_excel(self):
        if not self.excel_rows:
            messagebox.showerror("Erro", "Nenhum Excel importado.")
            return

        self.save_config()
        items = []
        for r in self.excel_rows:
            items.append({
                "chave": r.get("chave", ""),
                "email": r.get("email", ""),
                "assunto": r.get("assunto", ""),
                "corpo": r.get("corpo", ""),
                "origem": "excel",
            })
        self.run_batch_items(items)

    def run_batch_items(self, items: list[dict]):
        do_email = self.send_email_each.get()
        if do_email and not self.smtp_ready():
            messagebox.showerror("Erro", "Envio por e-mail marcado, mas SMTP/To padrão não está completo na Configuração.")
            return

        total = len(items)
        item_by_key = {item["chave"]: item for item in items}
        ok = 0
        falhas = []
        resultados = []

        self.progress.set(0.0)
        self.pb.set(0)
        self.log("=" * 110)
        self.log(f"Iniciando lote: {total} item(ns) | Pausa={self.pause_between_items.get()}s | Reprocess={self.reprocess_rounds.get()} rodada(s)")

        for i, item in enumerate(items, start=1):
            chave = item["chave"]
            origem = item.get("origem", "")

            self.log(f"[{i}/{total}] Processando chave: {chave}")
            pdf, err = self.baixar_pdf(chave)

            email_to = ""
            if pdf:
                ok += 1
                if do_email:
                    try:
                        email_to = self.send_email_for_pdf(chave=chave, pdf=pdf, email=item.get("email", ""), assunto=item.get("assunto", ""), corpo=item.get("corpo", ""))
                    except Exception as e:
                        self.log(f"  [EMAIL] ERRO: {e}")

                resultados.append({"chave": chave, "status": "OK", "pdf_path": str(pdf), "email_to": email_to, "erro": "", "origem": origem})
            else:
                falhas.append(chave)
                resultados.append({"chave": chave, "status": "FALHA", "pdf_path": "", "email_to": "", "erro": err, "origem": origem})
                self.log(f"  FALHA → {err}")

            time.sleep(float(self.pause_between_items.get()))
            ratio = i / total if total else 1
            self.progress.set(ratio * 100.0)
            self.pb.set(ratio)

            n = int(self.cooldown_every_n.get())
            if n > 0 and i % n == 0:
                time.sleep(int(self.cooldown_seconds.get()))

        rounds = int(self.reprocess_rounds.get())
        if falhas and rounds > 0:
            self.log("-" * 110)
            self.log(f"Entrando em reprocessamento: {len(falhas)} falha(s)")

            restantes = falhas[:]
            for rodada in range(1, rounds + 1):
                if not restantes:
                    break

                self.log(f"=== Rodada {rodada}/{rounds} | Restantes: {len(restantes)} ===")
                time.sleep(int(self.cooldown_between_rounds.get()))

                novas_restantes = []
                for j, chave in enumerate(restantes, start=1):
                    self.log(f"[RETRY {rodada}] ({j}/{len(restantes)}) {chave}")
                    pdf, err = self.baixar_pdf(chave)

                    if pdf:
                        ok += 1
                        for row in resultados:
                            if row["chave"] == chave and row["status"] == "FALHA":
                                row["status"] = "RECUPERADO"
                                row["pdf_path"] = str(pdf)
                                row["erro"] = ""
                                break

                        if do_email:
                            try:
                                original = item_by_key.get(chave, {})
                                email_to = self.send_email_for_pdf(
                                    chave=chave,
                                    pdf=pdf,
                                    email=original.get("email", ""),
                                    assunto=original.get("assunto", ""),
                                    corpo=original.get("corpo", ""),
                                )
                                for row in resultados:
                                    if row["chave"] == chave and row["status"] == "RECUPERADO":
                                        row["email_to"] = email_to
                                        break
                            except Exception as e:
                                self.log(f"  [EMAIL] ERRO: {e}")

                        self.log("  RECUPERADO ✅")
                    else:
                        novas_restantes.append(chave)
                        for row in resultados:
                            if row["chave"] == chave and row["status"] == "FALHA":
                                row["erro"] = err
                                break
                        self.log(f"  AINDA FALHA → {err}")

                    if int(self.cooldown_every_n.get()) > 0 and j % int(self.cooldown_every_n.get()) == 0:
                        time.sleep(int(self.cooldown_seconds.get()))
                    else:
                        time.sleep(float(self.pause_between_items.get()))

                restantes = novas_restantes

            falhas_finais = restantes[:]
        else:
            falhas_finais = falhas[:]

        out_dir = self._out_dir()
        resultado_path = salvar_resultado_xlsx(resultados, out_dir)
        self.log("-" * 110)
        self.log(f"Relatório salvo: {resultado_path}")

        falhas_path = None
        if falhas_finais:
            falhas_path = salvar_falhas_txt(falhas_finais, out_dir)
            self.log(f"Falhas finais salvas: {falhas_path}")

        self.log("-" * 110)
        self.log(f"Fim do lote. Sucesso total: {ok} | Falhas finais: {len(falhas_finais)}")
        if falhas_path:
            messagebox.showwarning(
                "Concluído (com falhas)",
                f"Lote finalizado.\n\nSucesso: {ok}\nFalhas finais: {len(falhas_finais)}\n\nRelatório: {resultado_path}\nFalhas: {falhas_path}",
            )
        else:
            messagebox.showinfo("Concluído", f"Lote finalizado com sucesso!\n\nSucesso: {ok}\n\nRelatório: {resultado_path}")


if __name__ == "__main__":
    try:
        from ctypes import windll

        windll.shcore.SetProcessDpiAwareness(1)
    except Exception:
        pass

    App().mainloop()
